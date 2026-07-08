import os
import sys
import re
import shutil
import threading
import subprocess
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.styles import Font
import win32com.client
import json

# ================== CHEMINS ==================

def get_app_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

APP_DIR      = get_app_dir()
DOSSIER_TEMP = os.path.join(APP_DIR, "temp_excel_print")
CONFIG_FILE  = os.path.join(APP_DIR, "prefixes.json")
MAX_RETRY    = 3

os.makedirs(DOSSIER_TEMP, exist_ok=True)

# ================== PREFIXES JSON ==================

DEFAULT_PREFIXES = {
    "CAGIP MoWER3"   : "GIM1P",
    "CASA MoWER3"    : "CASM1P",
    "CAGIP RedAlpha" : "GIFRG1L",
    "CALF MoWER3"    : "CALFM1P"
}

def charger_prefixes():
    if not os.path.exists(CONFIG_FILE):
        sauvegarder_prefixes(DEFAULT_PREFIXES)
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def sauvegarder_prefixes(data):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

PREFIXES = charger_prefixes()

# ================== EXCEL ==================

CELLULE_GIP     = "C37"
CELLULE_MACHINE = "B38"
CELLULE_HOST    = "C38"

def trouver_modele_excel():
    for f in os.listdir(APP_DIR):
        if f.lower().endswith(".xlsx"):
            return os.path.join(APP_DIR, f)
    return ""

def get_principal_cell(ws, cell):
    for merge in ws.merged_cells.ranges:
        if cell in merge:
            return merge.min_row, merge.min_col
    return ws[cell].row, ws[cell].column

def set_cell_value(ws, cell, value, size=18, bold=False):
    r, c = get_principal_cell(ws, cell)
    cobj = ws.cell(row=r, column=c)
    cobj.value = value
    cobj.font = Font(size=size, bold=bold)

# ================== IMPRESSION ==================

def lancer_impression_thread():
    threading.Thread(target=imprimer, daemon=True).start()

def imprimer():
    modele_excel = entry_fichier.get()
    type_machine = combo_machine.get()
    gip_list     = [g.strip() for g in text_gip.get("1.0", tk.END).splitlines() if g.strip()]

    if not os.path.isfile(modele_excel):
        messagebox.showerror("Erreur", "Fichier Excel introuvable.")
        return

    if type_machine not in PREFIXES:
        messagebox.showerror("Erreur", "Entité invalide.")
        return

    if use_manual_host.get():
        host_list = [h.strip() for h in text_host.get("1.0", tk.END).splitlines() if h.strip()]
        if len(host_list) != len(gip_list):
            messagebox.showerror("Erreur", "Nombre de HOSTNAME ≠ nombre de GIP.")
            return
    else:
        host_list = None

    prefix_base      = PREFIXES[type_machine]
    total            = len(gip_list)
    progress["maximum"] = total
    progress["value"]   = 0
    erreurs = []

    # ✅ Une seule instance Excel isolée (DispatchEx) ouverte pour toute la boucle,
    # fermée uniquement à la fin (ou en cas de plantage). Ne touche jamais aux
    # fichiers Excel déjà ouverts par ailleurs, mais évite de payer le coût de
    # démarrage/fermeture d'Excel à chaque GIP.
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible       = False
    excel.DisplayAlerts = False

    try:
        for i, gip in enumerate(gip_list, start=1):
            chiffres = ''.join(re.findall(r'\d+', gip))
            host     = host_list[i - 1] if host_list else f"{prefix_base}{chiffres}"

            succes      = False
            fichier_temp = None

            for tentative in range(1, MAX_RETRY + 1):
                wb_excel = None
                try:
                    fichier_temp = os.path.join(DOSSIER_TEMP, f"{host}.xlsx")
                    shutil.copy(modele_excel, fichier_temp)

                    # Remplissage via openpyxl
                    wb = load_workbook(fichier_temp)
                    ws = wb.active
                    set_cell_value(ws, CELLULE_MACHINE, type_machine, 18)
                    set_cell_value(ws, CELLULE_HOST,    host,         18)
                    set_cell_value(ws, CELLULE_GIP,     gip,          18)
                    wb.save(fichier_temp)
                    wb.close()

                    wb_excel = excel.Workbooks.Open(os.path.abspath(fichier_temp))
                    ws_excel = wb_excel.Worksheets(1)

                    ws_excel.PrintOut()
                    wb_excel.Close(False)
                    wb_excel = None

                    os.remove(fichier_temp)
                    fichier_temp = None

                    succes = True
                    break

                except Exception as e:
                    print(f"Erreur {host} (tentative {tentative}) : {e}")
                    try:
                        if wb_excel is not None:
                            wb_excel.Close(False)
                    except Exception:
                        pass
                    if fichier_temp and os.path.exists(fichier_temp):
                        try:
                            os.remove(fichier_temp)
                        except Exception:
                            pass
                    # Si l'instance Excel elle-même a planté (ex: RPC_E_DISCONNECTED),
                    # on en relance une nouvelle pour ne pas bloquer tout le lot.
                    try:
                        _ = excel.Visible
                    except Exception:
                        try:
                            excel.Quit()
                        except Exception:
                            pass
                        excel = win32com.client.DispatchEx("Excel.Application")
                        excel.Visible       = False
                        excel.DisplayAlerts = False

            if not succes:
                erreurs.append(host)

            progress["value"] = i
            progress_label.config(text=f"{i} / {total}")
            root.update_idletasks()

    finally:
        # ✅ Ferme l'instance Excel isolée une seule fois, à la toute fin du lot
        try:
            excel.Quit()
        except Exception:
            pass

    if erreurs:
        messagebox.showwarning("Terminé avec erreurs",
                               f"Échecs d'impression :\n{', '.join(erreurs)}")
    else:
        messagebox.showinfo("Terminé", "Toutes les impressions ont réussi ✅")

# ================== PREFIXES UI ==================

def ouvrir_prefixes_json():
    if not os.path.exists(CONFIG_FILE):
        sauvegarder_prefixes(DEFAULT_PREFIXES)
    subprocess.Popen(["notepad.exe", CONFIG_FILE])

def recharger_prefixes():
    global PREFIXES
    try:
        PREFIXES = charger_prefixes()
        menu = option_menu["menu"]
        menu.delete(0, "end")
        for entite in PREFIXES.keys():
            menu.add_command(label=entite, command=tk._setit(combo_machine, entite))
        messagebox.showinfo("OK", "Préfixes mis à jour ✅")
    except Exception as e:
        messagebox.showerror("Erreur JSON", str(e))

# ================== UI PRINCIPALE ==================

# ---- Palette et polices "standard entreprise" ----
COULEUR_FOND       = "#F4F6F8"
COULEUR_PANNEAU    = "#FFFFFF"
COULEUR_PRIMAIRE   = "#1F3B57"   # bleu marine corporate
COULEUR_ACCENT     = "#2E7D32"   # vert pour l'action principale
COULEUR_ACCENT_HVR = "#256428"
COULEUR_TEXTE      = "#22292F"
COULEUR_TEXTE_MUT  = "#5A6672"
COULEUR_BORDURE    = "#D5DBE1"

POLICE_BASE   = ("Segoe UI", 10)
POLICE_LABEL  = ("Segoe UI", 10, "bold")
POLICE_TITRE  = ("Segoe UI", 16, "bold")
POLICE_SOUS   = ("Segoe UI", 9)
POLICE_BOUTON = ("Segoe UI", 10, "bold")

root = tk.Tk()
root.title("Impression Excel Automatique")
root.configure(bg=COULEUR_FOND)
root.minsize(920, 560)

style = ttk.Style(root)
try:
    style.theme_use("clam")
except tk.TclError:
    pass

style.configure("TProgressbar", troughcolor=COULEUR_BORDURE,
                background=COULEUR_ACCENT, thickness=14)

# ---- En-tête ----
entete = tk.Frame(root, bg=COULEUR_PRIMAIRE, height=64)
entete.pack(fill="x", side="top")
entete.pack_propagate(False)

tk.Label(entete, text="Impression Excel Automatique",
         font=POLICE_TITRE, bg=COULEUR_PRIMAIRE, fg="white").pack(
    side="left", padx=20)
tk.Label(entete, text="Génération et impression en masse des fiches machine",
         font=POLICE_SOUS, bg=COULEUR_PRIMAIRE, fg="#D8E2EC").pack(
    side="left", padx=(0, 20), pady=(6, 0))

# ---- Corps principal ----
corps = tk.Frame(root, bg=COULEUR_FOND)
corps.pack(fill="both", expand=True, padx=20, pady=16)

def creer_panneau(parent, titre):
    panneau = tk.Frame(parent, bg=COULEUR_PANNEAU, highlightbackground=COULEUR_BORDURE,
                        highlightthickness=1, bd=0)
    tk.Label(panneau, text=titre, font=POLICE_LABEL,
             bg=COULEUR_PANNEAU, fg=COULEUR_PRIMAIRE).grid(
        row=0, column=0, columnspan=4, sticky="w", padx=14, pady=(12, 6))
    return panneau

# --- Panneau configuration (fichier + type de machine) ---
panneau_config = creer_panneau(corps, "Configuration")
panneau_config.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 12))
corps.grid_columnconfigure(0, weight=3)
corps.grid_columnconfigure(1, weight=2)

tk.Label(panneau_config, text="Fichier Excel modèle :", font=POLICE_BASE,
         bg=COULEUR_PANNEAU, fg=COULEUR_TEXTE).grid(
    row=1, column=0, sticky="w", padx=14, pady=6)
entry_fichier = tk.Entry(panneau_config, width=48, font=POLICE_BASE,
                          relief="solid", bd=1, highlightthickness=0)
entry_fichier.grid(row=1, column=1, padx=5, pady=6, sticky="w")

def parcourir_fichier():
    chemin = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
    if chemin:
        entry_fichier.delete(0, tk.END)
        entry_fichier.insert(0, chemin)

tk.Button(panneau_config, text="Parcourir", command=parcourir_fichier,
          font=POLICE_BASE, bg="#EAEFF3", fg=COULEUR_TEXTE, relief="flat",
          activebackground=COULEUR_BORDURE, padx=10).grid(
    row=1, column=2, padx=5, pady=6)
entry_fichier.insert(0, trouver_modele_excel())

tk.Label(panneau_config, text="Type de machine :", font=POLICE_BASE,
         bg=COULEUR_PANNEAU, fg=COULEUR_TEXTE).grid(
    row=2, column=0, sticky="w", padx=14, pady=(6, 14))
combo_machine = tk.StringVar(value="Sélectionner")
option_menu   = tk.OptionMenu(panneau_config, combo_machine, *PREFIXES.keys())
option_menu.configure(font=POLICE_BASE, bg="white", relief="solid", bd=1,
                       highlightthickness=0, activebackground=COULEUR_BORDURE)
option_menu["menu"].configure(font=POLICE_BASE)
option_menu.grid(row=2, column=1, sticky="w", padx=5, pady=(6, 14))

# --- Panneau saisie GIP / Hostname ---
panneau_saisie = creer_panneau(corps, "Saisie des données")
panneau_saisie.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(0, 12))
corps.grid_rowconfigure(1, weight=1)

tk.Label(panneau_saisie, text="Numéros GIP (1 par ligne) :", font=POLICE_BASE,
         bg=COULEUR_PANNEAU, fg=COULEUR_TEXTE).grid(
    row=1, column=0, sticky="w", padx=14, pady=(0, 4))
text_gip = tk.Text(panneau_saisie, width=52, height=14, font=POLICE_BASE,
                    relief="solid", bd=1, highlightthickness=0)
text_gip.grid(row=2, column=0, padx=14, pady=(0, 14), sticky="nsew")

use_manual_host = tk.BooleanVar()
tk.Checkbutton(panneau_saisie, text="Utiliser hostname manuel", variable=use_manual_host,
               font=POLICE_BASE, bg=COULEUR_PANNEAU, fg=COULEUR_TEXTE,
               activebackground=COULEUR_PANNEAU, selectcolor="white").grid(
    row=1, column=1, sticky="w", padx=(6, 14), pady=(0, 4))
tk.Label(panneau_saisie, text="", font=POLICE_BASE,
         bg=COULEUR_PANNEAU, fg=COULEUR_TEXTE_MUT).grid(
    row=1, column=1, sticky="e", padx=(0, 130))
text_host = tk.Text(panneau_saisie, width=32, height=14, font=POLICE_BASE,
                     relief="solid", bd=1, highlightthickness=0)
text_host.grid(row=2, column=1, padx=(6, 14), pady=(0, 14), sticky="nsew")

panneau_saisie.grid_columnconfigure(0, weight=3)
panneau_saisie.grid_columnconfigure(1, weight=2)
panneau_saisie.grid_rowconfigure(2, weight=1)

# --- Panneau action / progression ---
panneau_action = tk.Frame(corps, bg=COULEUR_PANNEAU, highlightbackground=COULEUR_BORDURE,
                           highlightthickness=1)
panneau_action.grid(row=2, column=0, columnspan=2, sticky="ew")

btn_lancer = tk.Button(panneau_action, text="▶  LANCER L'IMPRESSION",
                        command=lancer_impression_thread,
                        font=POLICE_BOUTON, bg=COULEUR_ACCENT, fg="white",
                        activebackground=COULEUR_ACCENT_HVR, activeforeground="white",
                        relief="flat", padx=18, pady=8, cursor="hand2")
btn_lancer.grid(row=0, column=0, padx=14, pady=14)

progress = ttk.Progressbar(panneau_action, length=420, style="TProgressbar")
progress.grid(row=0, column=1, padx=10, pady=14, sticky="ew")
progress_label = tk.Label(panneau_action, text="0 / 0", font=POLICE_BASE,
                           bg=COULEUR_PANNEAU, fg=COULEUR_TEXTE_MUT)
progress_label.grid(row=0, column=2, padx=(0, 14))
panneau_action.grid_columnconfigure(1, weight=1)

# --- Barre d'outils préfixes ---
barre_outils = tk.Frame(corps, bg=COULEUR_FOND)
barre_outils.grid(row=3, column=0, columnspan=2, sticky="w", pady=(12, 0))

def bouton_secondaire(parent, texte, commande):
    return tk.Button(parent, text=texte, command=commande,
                      font=POLICE_BASE, bg="#EAEFF3", fg=COULEUR_TEXTE,
                      activebackground=COULEUR_BORDURE, relief="flat",
                      padx=10, pady=4, cursor="hand2")

bouton_secondaire(barre_outils, "📝 Modifier les préfixes", ouvrir_prefixes_json).pack(
    side="left", padx=(0, 8))
bouton_secondaire(barre_outils, "🔄 Recharger les préfixes", recharger_prefixes).pack(
    side="left")

# ---- Pied de page ----
pied_de_page = tk.Frame(root, bg=COULEUR_PRIMAIRE, height=30)
pied_de_page.pack(fill="x", side="bottom")
pied_de_page.pack_propagate(False)
tk.Label(pied_de_page, text="Créé par Cristian", font=("Segoe UI", 9),
         bg=COULEUR_PRIMAIRE, fg="#D8E2EC").pack(side="right", padx=16, pady=5)

root.mainloop()
