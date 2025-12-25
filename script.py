import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.styles import Font
import win32com.client
import win32print
import threading
import shutil
import os
import re
import pythoncom
# ================= CONFIG =================
MAX_RETRY = 3   # nombre de tentatives par GIP
count=0
DOSSIER_TEMP = os.path.join(os.path.expanduser("~"), "temp_excel_print")
os.makedirs(DOSSIER_TEMP, exist_ok=True)

PREFIXES = {
    "CAGIP MoWER3": "GIM1P",
    "CASA MoWER3": "CASM1P",
    "CAGIP RedAlpha": "GIFRG1L",
    "BFB MoWER3": "SELM1P",
    "CAIMMO MOWER3": "CAIM1P",
    "CALF MoWER3": "CALFM1P",
    "CAPS MoWER3": "CAPSM1P",
    "CAGIP RedAlphaUK": "GIUKG1L"
}

CELLULE_GIP = "B9"
CELLULE_MACHINE = "B36"
CELLULE_HOST = "C36"
CELLULE_MASTER = "B12"
CELLULE_ENTITER = "B17"

# ================= EXCEL UTILS =================
def trouver_modele_excel():
    dossier_app = os.path.dirname(os.path.abspath(__file__))
    for f in os.listdir(dossier_app):
        if f.lower().endswith(".xlsx"):
            return os.path.join(dossier_app, f)
    return None

def get_principal_cell(ws, cell):
    for merge in ws.merged_cells.ranges:
        if cell in merge:
            return merge.min_row, merge.min_col
    return ws[cell].row, ws[cell].column

def set_cell_value(ws, cell, value, size=18, bold=False):
    r, c = get_principal_cell(ws, cell)
    cobj = ws.cell(row=r, column=c)
    cobj.value = value
    cobj.font = Font(size=size, bold=bold)


# ================= THREAD =================

def lancer_impression_thread():
    threading.Thread(target=imprimer, daemon=True).start()

# ================= IMPRESSION =================

def imprimer():
    modele_excel = entry_fichier.get()
    type_machine = combo_machine.get()
    gip_list = [g.strip() for g in text_gip.get("1.0", tk.END).splitlines() if g.strip()]

    if not os.path.isfile(modele_excel):
        messagebox.showerror("Erreur", "Fichier Excel introuvable.")
        return

    prefix_base = PREFIXES.get(type_machine)
    if not prefix_base:
        messagebox.showerror("Erreur", "Type de machine invalide.")
        return

    total = len(gip_list)
    progress["maximum"] = total
    progress["value"] = 0

    erreurs = []

    for i, gip in enumerate(gip_list, start=1):
        chiffres = ''.join(re.findall(r'\d+', gip))
        prefix = prefix_base

        if type_machine == "CALF MoWER3" and len(chiffres) == 6:
            prefix = "CALFM1P0"

        host = f"{prefix}{chiffres}"

        succes = False

        for tentative in range(1, MAX_RETRY + 1):
            try:
                fichier_temp = os.path.join(
                    DOSSIER_TEMP, f"{host}_temp.xlsx"
                )

                #  Copie
                shutil.copy(modele_excel, fichier_temp)

                #  Écriture openpyxl
                wb = load_workbook(fichier_temp)
                ws = wb.active

                set_cell_value(ws, CELLULE_MACHINE, type_machine, 18)
                set_cell_value(ws, CELLULE_HOST, host, 15)
                set_cell_value(ws, CELLULE_GIP, gip, 33)

                entite, master = type_machine.split(" ")
                set_cell_value(ws, CELLULE_MASTER, master, 40)
                set_cell_value(ws, CELLULE_ENTITER, entite, 40)

                wb.save(fichier_temp)
                wb.close()

                # Impression Excel invisible
                excel = win32com.client.Dispatch("Excel.Application")
                wb_excel = excel.Workbooks.Open(os.path.abspath(fichier_temp))
                wb_excel.Worksheets(1).PrintOut()
                wb_excel.Close(False)
                excel.Quit()
                os.remove(fichier_temp)

                succes = True
                break  # on sort des retries

            except Exception as e:
                print(f"❌ Erreur {host} (tentative {tentative}) : {e}")

                try:
                    excel.Quit()
                except:
                    pass

        if not succes:
            erreurs.append(host)

        progress["value"] = i
        root.update_idletasks()
        progress_label.config(text=f"{i} / {total}")
        root.update_idletasks()

    if erreurs:
        messagebox.showwarning(
            "Terminé avec erreurs",
            f"Impressions réussies : {total - len(erreurs)}\n"
            f"Échecs : {len(erreurs)}\n\n"
            f"{', '.join(erreurs)}"
        )
    else:
        messagebox.showinfo("Terminé", f"{total} impressions réussies.")
# ================= UI =================

def choisir_fichier():
    f = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if f:
        entry_fichier.delete(0, tk.END)
        entry_fichier.insert(0, f)

root = tk.Tk()
root.title("Impression Excel Automatique ")

tk.Label(root, text="Fichier Excel modèle :").grid(row=0, column=0, sticky="w")
entry_fichier = tk.Entry(root, width=50)
entry_fichier.grid(row=0, column=1)
tk.Button(root, text="Parcourir", command=choisir_fichier).grid(row=0, column=2)

modele_excel = trouver_modele_excel()
if modele_excel:
    entry_fichier.insert(0, modele_excel)

tk.Label(root, text="Type de machine :").grid(row=1, column=0, sticky="w")
combo_machine = tk.StringVar(value="Choisissez l'entité")
tk.OptionMenu(root, combo_machine, *PREFIXES.keys()).grid(row=1, column=1, sticky="w")

tk.Label(root, text="Numéros GIP (1 par ligne) :").grid(row=2, column=0, sticky="nw")
text_gip = tk.Text(root, width=70, height=20)
text_gip.grid(row=2, column=1, columnspan=2)

tk.Button(
    root,
    text="LANCER IMPRESSION",
    command=lancer_impression_thread,
    bg="green",
    fg="white"
).grid(row=3, column=1, pady=10)

progress = ttk.Progressbar(root, length=400, mode="determinate")
progress.grid(row=4, column=1, pady=10)

progress_label = tk.Label(root, text="0 / 0")
progress_label.grid(row=4, column=2, padx=10)
root.mainloop()
